import openpyxl
import os
from PIL import Image, ImageDraw, ImageFont

workbook = openpyxl.load_workbook("Nomes.xlsx")
sheet = workbook["Certificados"]

corTexto = (0, 0, 0)

nomesAlunos = [
    sheet.cell(row=row, column=1).value
    for row in range(2, sheet.max_row + 1)
]

nomesPalestrantes = [
    sheet.cell(row=row, column=2).value
    for row in range(2, sheet.max_row + 1)
]

os.makedirs("certificados prontos", exist_ok=True)

for aluno, palestrante in zip(nomesAlunos, nomesPalestrantes):

    caminho_pasta = f"certificados prontos/{palestrante}"
    os.makedirs(caminho_pasta, exist_ok=True)

    caminhoModelo = f"modelos certificados/{palestrante}.jpg"

    if os.path.exists(caminhoModelo):
        tamanhoFonte = 180
        if len(aluno)>=25:
            tamanhoFonte = 150
            y= 850
            x=1100
        elif len(aluno)<=18:
            x= 1320
            y= 820
        else:
            y = 820
            x= 1200
        fonte = ImageFont.truetype("Baliana.ttf", tamanhoFonte)
        certificado = Image.open(caminhoModelo)
        desenho = ImageDraw.Draw(certificado)



        desenho.text((x, y), aluno.title(), fill=corTexto, font=fonte)

        caminhoCertificado = f"{caminho_pasta}/{aluno}.jpg"
        certificado.save(caminhoCertificado)


